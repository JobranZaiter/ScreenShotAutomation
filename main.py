import os
import time
from io import BytesIO
from selenium import webdriver
from selenium.common import StaleElementReferenceException, TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import logging
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from urllib.parse import urlparse, urljoin
from PIL import Image as PILImage


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)

CHATGPT_API_KEY = 'xxxxxxx'


def is_image_url(url):
    image_extensions = ('.jpeg', '.jpg', '.png', '.gif', '.bmp', '.tiff', '.svg', '.pdf')
    return url.endswith(image_extensions)


def ask_chatgpt(prompt):
    logging.info('Sending prompt')
    chatgpt_api_url = "https://api.openai.com/v1/chat/completions"
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {CHATGPT_API_KEY}"}
    data = {
        "model": "gpt-4o",
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": 2000,
        "temperature": 0.7
    }
    logging.debug(f"Sending prompt to OpenAI: {prompt}")
    response = requests.post(chatgpt_api_url, headers=headers, json=data)
    response.raise_for_status()
    logging.debug(f"Response from OpenAI: {response.json()}")
    return response.json()


def driver_start():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    chrome_driver_path = "C:/Users/User/Downloads/Programs/chromedriver-win64/chromedriver.exe"
    service = Service(executable_path=chrome_driver_path)
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def get_links(url):
    web_driver = driver_start()
    domain = urlparse(url).netloc
    stack = [url]
    internal_links = set()
    crawled = set()
    appended = set()
    max_retries = 5
    delay = 5

    while stack:
        retries = 0
        logging.debug('Entering')
        current_url = stack.pop()
        logging.info(f'Processing: {current_url}')
        while retries < max_retries:
            try:
                web_driver.get(current_url)
                logging.debug(f'Got URL {current_url}')
                break
            except Exception as e:
                print(f'Error while fetching url: {current_url}\n Exception: {e}')
                retries += 1
                if retries < max_retries:
                    print(f'Retrying to fetch url: {current_url}')
                    time.sleep(delay)
                else:
                    print(f'Failed to fetch URL:{current_url}')

        if current_url in crawled:
            continue
        internal_links.add(current_url)
        try:
            links = WebDriverWait(web_driver, 10).until(
                EC.presence_of_all_elements_located((By.TAG_NAME, 'a'))
            )
            logging.debug(f"Found Links for {current_url}")
            for link in links:
                if not is_image_url(link.get_attribute('href')):
                    next_url = urljoin(current_url, link.get_attribute('href'))
                    if urlparse(
                            next_url).netloc == domain and next_url not in internal_links and next_url not in appended:
                        stack.append(next_url)
                        appended.add(next_url)
        except Exception as e:
            print(f'Exception while fetching links for url: {current_url}\n Exception: {e}')

        crawled.add(current_url)
    web_driver.close()

    return list(internal_links)


def get_links2(url):
    web_driver = driver_start()
    domain = urlparse(url).netloc
    stack = [url]
    internal_links = set()
    crawled = set()
    max_retries = 5
    delay = 5

    try:
        while stack:
            retries = 0
            current_url = stack.pop()

            if current_url in crawled:
                continue

            while retries < max_retries:
                try:
                    web_driver.get(current_url)
                    logging.debug(f'Got URL {current_url}')
                    break
                except Exception as e:
                    print(f'Error while fetching url: {current_url}\n Exception: {e}')
                    retries += 1
                    if retries < max_retries:
                        print(f'Retrying to fetch url: {current_url}')
                        time.sleep(delay)
                    else:
                        print(f'Failed to fetch URL: {current_url}')

            crawled.add(current_url)

            try:
                links = WebDriverWait(web_driver, 10).until(
                    EC.presence_of_all_elements_located((By.TAG_NAME, 'a'))
                )
                logging.debug(f"Found Links for {current_url}")
                for link in links:
                    href = link.get_attribute('href')
                    if href and not is_image_url(href):
                        next_url = urljoin(current_url, href)
                        next_url = urlparse(next_url)._replace(fragment='').geturl()
                        if urlparse(next_url).netloc == domain and next_url not in internal_links:
                            stack.append(next_url)
                            internal_links.add(next_url)
            except Exception as e:
                print(f'Exception while fetching links for url: {current_url}\n Exception: {e}')
    finally:
        web_driver.quit()

    return list(internal_links)


def scrape_images(urls):
    img_data = []
    processed_images = set()
    driver = driver_start()

    for url in urls:

        try:
            logging.info(f'Attempting to get URL: {url}')
            driver.get(url)
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//body'))
            )
            logging.info(f'Successfully got URL: {url}')
        except Exception as e:
            logging.error(f'Failed to get URL: {url}\nException: {e}')
            continue

        scroll_down(driver)

        try:
            logging.info(f'Attempting to get initial images for url: {url}')
            initial_images = WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located((By.TAG_NAME, 'img'))
            )
            img_data.extend(process_image_data(initial_images, url, processed_images))
            logging.info(f'Successfully scraped initial images: {len(initial_images)}')
        except Exception as e:
            logging.error(f'Could not process images: {e}')

        try:
            logging.info(f'Attempting to get clickable elements for url: {url}')
            elements = get_clickable_elements(driver, url)
            logging.info(f'Successfully got {len(elements)} clickable elements')
        except Exception as e:
            logging.error('Could not get clickable elements')
            continue

        processed_elements = []
        for attributes in elements:
            process_element(driver, attributes, img_data, processed_elements, processed_images, url)

    driver.quit()
    logging.info(f'Successfully scraped images: {len(img_data)}')
    return img_data


def process_image_data(images, url, processed_images):
    image_data = []
    for image in images:
        try:
            source = image.get_attribute('src')
            if source:
                source = urljoin(url, source)
                if source not in processed_images:
                    alt = image.get_attribute('alt') or 'No Alt'
                    image_data.append((source, url, alt))
                    processed_images.add(source)
                else:
                    logging.warning(f'Duplicate source for image on URL: {url} - {source}')
            else:
                logging.warning(f'No source for image on URL: {url}')
        except Exception as e:
            logging.error(f'Error processing image information on URL: {url}\nException: {e}')
    return image_data


def scroll_down(driver):
    try:
        height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(5)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == height:
                logging.info('Reached the bottom of the page.')
                break
            height = new_height
    except Exception as e:
        logging.error(f'Error during scroll down: {e}')


def get_clickable_elements(driver, url):
    elements = []
    try:
        scroll_down(driver)
        clickable_elements = driver.find_elements(By.CSS_SELECTOR, "*[onclick], button, [role='button'], a[href]")
        elements = [get_element_attributes(driver, element) for element in clickable_elements if
                    urlparse(element.get_attribute('href')).netloc == url or element.get_attribute('href') is None]
        pointer_elements = driver.find_elements(By.CSS_SELECTOR, "[style='cursor: pointer']")
        elements += [get_element_attributes(driver, element) for element in pointer_elements if (
                urlparse(element.get_attribute('href')).netloc == url or element.get_attribute(
            'href') is None) and element not in clickable_elements]
        logging.debug(f'Found {len(elements)} clickable elements.')
    except Exception as e:
        logging.error(f'Error getting clickable elements: {e}')
    return elements


def get_element_attributes(driver, element):
    attributes = driver.execute_script("""
    function getElementAttributesAsDict(element) {
        var attributesDict = {};
        attributesDict["tag"] = element.tagName.toLowerCase();
        Array.from(element.attributes).forEach(attr => {
            attributesDict[attr.name] = attr.value;
        });
        return attributesDict;
    }
    return getElementAttributesAsDict(arguments[0]);
    """, element)
    return attributes


def find_element_by_attributes(driver, attributes):
    selector = f"{attributes.get('tag', '')}"
    for attr, value in attributes.items():
        if attr != 'tag' and value:
            selector += f"[{attr}='{value}']"

    try:
        web_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
        )
        return web_element
    except TimeoutException as e:
        logging.warning(f'Could not find element, timed out: {e}')
    except StaleElementReferenceException as e:
        logging.warning(f'Stale element reference: {e}')
    except Exception as e:
        logging.warning(f'Error while finding element by attributes: {e}')

    logging.info('No Web Element found')
    return None


def process_element(driver, attributes, img_data, elements, processed_images, url):
    retry_count = 3
    for _ in range(retry_count):
        try:
            element = find_element_by_attributes(driver, attributes)
            if element:
                logging.info('Scrolling elements into view')
                driver.execute_script("arguments[0].scrollIntoView(true)", element)
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(element)
                )
                logging.info('Successfully scrolled element into view')
                time.sleep(1)
                element.click()
                logging.info('Clicked element')
                WebDriverWait(driver, 10)
                logging.info('111111')
                new_url = driver.current_url
                if new_url != url:
                    logging.info(f'New url detected {new_url}')
                    driver.get(url)
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.TAG_NAME, 'body'))
                    )
                logging.info('Getting clickable elements')
                new_elements = get_clickable_elements(driver, url)
                for new_element in new_elements:
                    new_frozenset_element = frozenset(new_element.items())
                    if new_frozenset_element not in elements:
                        elements.append(new_frozenset_element)
                logging.info('Successfully got clickable elements')
                logging.info('Getting new images')
                new_images = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.TAG_NAME, 'img'))
                )
                img_data.extend(process_image_data(new_images, url, processed_images))
            break
        except StaleElementReferenceException:
            logging.warning(f'Stale element reference, retrying: {attributes}')
        except Exception as e:
            logging.error(f'Error processing element: {e}')
            break


def resize_image(path):
    max_width = 100
    max_height = 100
    try:
        with PILImage.open(path) as img:
            width, height = img.size
            scale = min(max_width / width, max_height / height)
            new_width = int(width * scale)
            new_height = int(height * scale)
            resized_img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            resized_img.save(path)
            return new_width, new_height
    except Exception as e:
        logging.error(f'Couldnt resize image with path: {path}')


def write_to_excel(image_arr, excel_filename, sheet_name):
    try:
        logging.info('Instantiating dataframe')
        img_df = pd.DataFrame(image_arr, columns=['Src', 'Url', 'Alt'])
        logging.info('Successfully created dataframe')
    except Exception as e:
        logging.error('Could not initialize dataframe')
        return

    try:
        logging.info('Instantiating Workbook and Sheet')
        if os.path.exists(excel_filename):
            logging.info('Filename exists')
            wb = load_workbook(excel_filename)
            if sheet_name in wb.sheetnames:
                logging.info('Worksheet exists')
                ws = wb[sheet_name]
            else:
                logging.info('Worksheet does not exist')
                ws = wb.create_sheet(sheet_name)
        else:
            logging.info('Workbook does not exist')
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
        logging.info('Successfully instantiated workbook and sheet')
    except Exception as e:
        logging.error('Error creating or loading workbook')
        return

    ws['A1'] = 'Image'
    ws['B1'] = 'Src'
    ws['C1'] = 'Url'
    ws['D1'] = 'Alt'

    remove_files = []
    for index, row in img_df.iterrows():
        img_filename = None
        try:
            logging.info(f'Attempting to download image from {row["Src"]}')
            img_url = row['Src']
            response = requests.get(img_url, stream=True)
            response.raise_for_status()
            img = PILImage.open(BytesIO(response.content))
            img_filename = f"{sheet_name}image{index + 1}.png"
            img.save(img_filename)
            remove_files.append(img_filename)
            logging.info(f'Successfully downloaded image with filename {img_filename}')
        except Exception as e:
            logging.error(f'Failed to download image: {e}')
            img_filename = None

        try:
            logging.info(f'Creating row with values {row["Src"]} {row["Url"]} {row["Alt"]}')
            if img_filename and os.path.exists(img_filename):
                new_width, new_height = resize_image(img_filename)
                img = Image(img_filename)
                cell = f'A{index + 2}'
                ws.add_image(img, cell)
                ws.column_dimensions[cell[0]].width = new_width / 7
                ws.row_dimensions[int(cell[1:])].height = new_height / 0.75
            else:
                ws[f'A{index + 2}'] = "Failed to download image"
            ws[f'B{index + 2}'] = row['Src']
            ws[f'C{index + 2}'] = row['Url']
            ws[f'D{index + 2}'] = row['Alt']
            logging.info('Successfully created row')
        except Exception as e:
            logging.error(f'Error writing row to Excel: {e}')
            continue

    try:
        logging.info(f'Saving workbook to {excel_filename}')
        wb.save(excel_filename)
    except PermissionError as e:
        logging.error(f'Permission denied while saving workbook: {e}')
    except Exception as e:
        logging.error(f'Error saving workbook: {e}')

    for file in remove_files:
        try:
            if os.path.exists(file):
                os.remove(file)
                logging.info(f'Successfully removed file : {file}')
            else:
                logging.error('Trying to remove a path that doesnt exist')
        except Exception as e:
            logging.error(f'Error removing file {file}')


def alt_generator(source, url, alt):
    try:
        response = requests.get(source, stream=True)
        response.raise_for_status()
        img = PILImage.open(BytesIO(response.content))
        img_filename = "target_image.png"
        img.save(img_filename)
        logging.info(f"Successfully downloaded image from {source}")
    except requests.exceptions.RequestException as e:
        logging.error(f"Couldn't fetch image from {source}\nException: {e}")
        return "Could not fetch the image"

    try:
        prompt = f"""
            You are given an image stored in the local file system with the name {img_filename}.
            Provide a alt tag for this image to optimize search engine results. Here are additional details:
            - Source URL: {source}
            - Original Alt Text: {alt}
            - Page URL: {url}
            Please generate a descriptive alternative text (alt) for the image.
            """

        chatgpt_response = ask_chatgpt(prompt)
        updated_alt_text = chatgpt_response.get("choices", [{}])[0].get("message", {}).get("content", "").strip()

        logging.info(f"Generated alt text: {updated_alt_text}")
        return updated_alt_text

    except Exception as e:
        logging.error(f"Error during alt text generation for {source}\nException: {e}")
        return {"Error generating alt text"}


def alt_writer(excel_filename, sheet_name, alt):
    try:
        logging.info('Instantiating Workbook and Sheet')
        wb = load_workbook(excel_filename)
        ws = wb[sheet_name]
        ws['E1'] = 'Updated Alt'
        logging.info('Successfully loaded workbook and sheet name')
    except Exception as e:
        logging.error(
            f'Error Couldnt instantiate workbook with filename: {excel_filename} and with sheet name: {sheet_name}\n{e} ')
        return
    try:
        logging.info('Loading info into dataframe')
        image_df = pd.read_excel(excel_filename, sheet_name=sheet_name)
        logging.info('Successfully loaded info into dataframe')
    except Exception as e:
        logging.error('Couldnt load info into dataframe')

    for index, row in image_df.iterrows():
        try:
            logging.info(f'Row {index + 1}: Src={row["Src"]}, Url={row["Url"]}, Alt={row["Alt"]}')
            ws[f'E{index + 2}'] = alt_generator(row['Src'], row['Url'], row['Alt'])
        except Exception as e:
            logging.error('Couldnt print row')
    try:
        logging.info(f'Saving workbook to {excel_filename}')
        wb.save(excel_filename)
    except PermissionError as e:
        logging.error(f'Permission denied while saving workbook: {e}')
    except Exception as e:
        logging.error(f'Error saving workbook: {e}')


def main():
    links = get_links2('https://pizzanini.no/')
    for link in links:
        print(link)
    scrape_images(links)


if __name__ == '__main__':
    main()
